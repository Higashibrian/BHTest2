#
import sys, os, openpyxl
from openpyxl import load_workbook
#For this script add the OR and FL folders to PATH env variable. This allows the FL and OR to be called as python modules.
#sys.path.append(self.path_to_file+'/FL')
#sys.path.append(self.path_to_file+'/OR')

#Naming convention for pages. Reference project name. Ex. NJFC 
import openpyxl,os,datetime
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Alignment, Font, PatternFill, Border, Side
from openpyxl.cell import Cell

#Create list of key pair values. Action1:(2,5) Action2 10,12
#Write metrics to a new sheet 
#Create "Error" function 

#results=BHTest.nametest(datafilename,"Product: Screening_App","Automated front end message validation","S:\QA\Projects\Screening")
					#(self, testname, 		title, 				description, 							path_to_file):
global debug
debug=False

# self.tccolumn="C"
# self.startrow=0
# self.statuscolumn="G"
# self.arcolumn="F"
# self.ercolumn="E"
# self.cmntcolumn="H"
# self.desccolumn="D"
# self.stepcolumn="B"

class nametest:
	def __init__(self, testname, title, description, path_to_file,column_names={}):
		self.path_to_file = path_to_file

		#print("path to file: "+self.path_to_file

		self.action_name=""
		self.current_action=0
		self.current_action_pass=0
		self.current_action_fail=0
		self.current_action_start = 0
		self.current_action_end = 0

		self.total_action_pass=0
		self.total_action_fail=0
		self.total_action_end=0

		self.total_pass=0
		self.total_fail=0

		self.date=datetime.datetime.now().strftime("%m_%d")
		self.datestring=datetime.datetime.now().strftime("%m%d_%H%M%S")
		try:
			os.mkdir(self.path_to_file+"/results")
		except:
			pass
		try:
			os.mkdir(self.path_to_file+"/results/"+self.date)
		except:
			pass
		try:
			os.mkdir(self.path_to_file+"/results/"+self.date+'/'+testname)
		except:
			pass


		self.tccolumn="C"
		self.startrow=0
		self.statuscolumn="G"
		self.arcolumn="F"
		self.ercolumn="E"
		self.cmntcolumn="H"
		self.desccolumn="D"
		self.stepcolumn="B"



		self.stepno=1
		self.currentrow=self.stepno+self.startrow

		self.testname = testname
		self.passno=0
		self.directory=self.path_to_file
		
		self.resultsbook=Workbook()

		self.bookname = self.directory+"/results/"+self.date+"/"+self.testname+'/'+self.datestring+".xlsx"
		try:
			self.resultsbook.save(filename = self.bookname)
		except:
			namecopy=self.bookname
			namecopy=namecopy.replace('//','/')
			namecopy=namecopy.replace('\\\\','/')
			namecopy=namecopy.replace('\\','/')
			dirlist=namecopy.split('/')
			makethisdir=''
			for item in dirlist:
				makethisdir+=item+'/'
				try:
					os.makedir(makethisdir)
				except:
					pass
		self.results = self.resultsbook.get_active_sheet()

		self.action_description_column_linebreak=131

		self.square_fill = PatternFill(start_color='AAAAAAAA',
		                   end_color='AAAAAAAA',
		                   fill_type='solid')

		self.bd = Side(style='thin', color="000000")
		self.document_font="Calisto MT"

		self.doc_font=Font(name = self.document_font, size =12)
		self.start_font=Font(name = self.document_font, bold=True, size = 18)
		self.start_fill = PatternFill(start_color='99999999', end_color='99999999',fill_type='solid')

		self.insert_title(title,description)
		self.print_titles(3)

	def insert_title(self,title, description):
		title = str(title)
		description = str(description)
		thin = Side(border_style="thin", color="000000")
		#double = Side(border_style="double", color="ff0000")
		border = Border(top=thin, left=thin, right=thin, bottom=thin)
		fill = PatternFill("solid", fgColor="DDDDDD")
		font = Font(name=self.document_font, size=18, b=True, color="000000")
		al = Alignment(horizontal="center", vertical="center", wrap_text=True)
		action_title=str(self.stepcolumn)+str(self.currentrow)

		self.color_square()
		self.style_range(self.results, "A"+str(self.currentrow)+':H'+str(self.currentrow), border, fill, font, al)
		self.results["A"+str(self.currentrow)]=title
		self.currentrow+=1

		#Description Row
		linebreak=int(self.action_description_column_linebreak)
		description_length=len(description)
		#print(description_length
		self.results.row_dimensions[self.currentrow].height=int(15*description_length/linebreak+15)


		self.color_square()
		thin = Side(border_style="thin", color="000000")
		#double = Side(border_style="double", color="ff0000")
		border = Border(top=thin, left=thin, right=thin)
		fill = PatternFill("solid", fgColor="FFFFFFFF")
		font = Font(size=12, b=False, color="000000", name=self.document_font)
		al = Alignment(horizontal="center", vertical="center", wrap_text=True)
		action_description=str(self.stepcolumn)+str((self.currentrow))
		self.style_range(self.results, "B"+str(self.currentrow)+':H'+str(self.currentrow), border, fill, font, al)
		self.results["B"+str(self.currentrow)]=description
		
		description_length=len(str(description))

		#Description is set
		self.currentrow+=1
		#self.results[]=title
		#self.results[action_description]=description
		self.stepno=1
		self.resultsbook.save(filename = self.bookname)
		self.currentrow+=1

		self.resultsbook.save(filename = self.bookname)
	def color_square(self):
		#print("A"+str(self.currentrow)
		self.results["A"+str(self.currentrow)].fill=self.square_fill
		self.results["A"+str(self.currentrow)].border=Border(right = self.bd)

	def print_titles(self,row):
		self.currentrow+=1
		self.color_square()
		stepcolumn=self.results[self.stepcolumn+str(row)]
		tccolumn=self.results[self.tccolumn+str(row)]
		statuscolumn=self.results[self.statuscolumn+str(row)]
		arcolumn=self.results[self.arcolumn+str(row)]
		ercolumn=self.results[self.ercolumn+str(row)]
		cmntcolumn=self.results[self.cmntcolumn+str(row)]
		desccolumn=self.results[self.desccolumn+str(row)]
		self.results[self.stepcolumn+str(row)]="Step"
		self.results[self.tccolumn+str(row)]="Testcase"
		self.results[self.statuscolumn+str(row)]="Status"
		self.results[self.arcolumn+str(row)]="Actual Result"
		self.results[self.ercolumn+str(row)]="Expected Result"
		self.results[self.cmntcolumn+str(row)]="Comments        "
		self.results[self.desccolumn+str(row)]="Description     "
#		print('print titles4')
		cell_list=[]
		cell_list.append(stepcolumn)
		cell_list.append(tccolumn)
		cell_list.append(statuscolumn)
		cell_list.append(arcolumn)
		cell_list.append(ercolumn)
		cell_list.append(cmntcolumn)
		cell_list.append(desccolumn)
		#print('print titles5')
		zoink=0
		for cell in cell_list:
			#print(zoink)
			zoink+=1
			cell.font=self.start_font
			#print('zoink.1')
			cell.fill=self.start_fill
			#print('zoink.2')
#			self.results.column_dimensions[cell.column].width=len(str(cell.value))*2.5
			#print('zoink.3')
			cell.border = Border(left = self.bd, right = self.bd, bottom = self.bd, top = self.bd)
			#print('end for')
	def comments_and_description(self,comments,description):
		self.results[self.cmntcolumn+str(self.currentrow)]=comments
		self.results[self.desccolumn+str(self.currentrow)]=description

	def actual_and_expected(self,actual,expected):
		self.results[self.arcolumn+str(self.currentrow)]=actual
		self.results[self.ercolumn+str(self.currentrow)]=expected

	def format_row(self,font):
		statuscolumn=self.results[self.statuscolumn+str(self.currentrow)]
		arcolumn=self.results[self.arcolumn+str(self.currentrow)]
		ercolumn=self.results[self.ercolumn+str(self.currentrow)]
		cmntcolumn=self.results[self.cmntcolumn+str(self.currentrow)]
		desccolumn=self.results[self.desccolumn+str(self.currentrow)]
		stepcolumn=self.results[self.stepcolumn+str(self.currentrow)]
		tccolumn=self.results[self.tccolumn+str(self.currentrow)]
		
		wrap_alignment = Alignment(wrap_text=True)

		cells_list=[]
		cells_list.append(statuscolumn)
		cells_list.append(arcolumn)
		cells_list.append(ercolumn)
		cells_list.append(cmntcolumn)
		cells_list.append(desccolumn)
		cells_list.append(tccolumn)
		cells_list.append(stepcolumn)

		for cell in cells_list:
			cell.font=font
			cell.alignment=wrap_alignment

	def step(self,description, comments):
		self.format_row(self.doc_font)
		self.color_square()
		self.comments_and_description(comments,description)
		self.results[self.tccolumn+str(self.currentrow)]=self.action_name
		self.results[self.stepcolumn+str(self.currentrow)]="step "+str(self.stepno)

		self.resultsbook.save(filename = self.bookname)
		self.stepno+=1
		self.currentrow+=1

	def step_pass(self, description, expected,actual,comments):
		try:
			self.format_row(self.doc_font)
			self.color_square()
			self.comments_and_description(comments,description)
			self.actual_and_expected(actual,expected)
			pass_fill = PatternFill(start_color='5500FF00',
			                   end_color='5500FF00',
			                   fill_type='solid')
			self.results[self.stepcolumn+str(self.currentrow)]="step "+str(self.stepno)
			self.results[self.tccolumn+str(self.currentrow)]=self.action_name
			self.results[self.statuscolumn+str(self.currentrow)]="Pass"
			self.results[self.statuscolumn+str(self.currentrow)].fill=pass_fill
			self.resultsbook.save(filename = self.bookname)
			self.stepno+=1
			self.currentrow+=1
			self.total_action_pass+=1
			self.total_pass+=1
			self.results['I6']=self.total_action_pass
			self.results['J6']=self.total_pass
			self.current_action_pass+=1
		except Exception as e:
			self.step_error(description, expected, actual,e)


	def step_fail(self,description, expected,actual,comments):
		try:
			self.format_row(self.doc_font)
			self.color_square()
			self.comments_and_description(comments,description)
			self.actual_and_expected(actual,expected)
			fail_fill = PatternFill(start_color='00FF0000',
			                   end_color='00FF0000',
			                   fill_type='solid')
			self.results[self.stepcolumn+str(self.currentrow)]="step "+str(self.stepno)
			self.results[self.tccolumn+str(self.currentrow)]=self.action_name
			self.results[self.statuscolumn+str(self.currentrow)]="Fail"
			self.results[self.statuscolumn+str(self.currentrow)].fill=fail_fill	
			self.resultsbook.save(filename = self.bookname)
			self.stepno+=1
			self.currentrow+=1
			self.total_action_fail+=1
			self.total_fail+=1
			self.results['I7']=self.total_action_fail
			self.results['J7']=self.total_fail
			self.current_action_fail+=1
		except Exception as e:
			self.step_error(description, expected, actual,e)	
	def step_error(self, description, expected,actual,comments):
		##print('bhtest error'
		try:
			##print("Error Message(BHTest): \nline: "+str(self.stepno)+"\n"+description+"\n"+str(comments)
			self.format_row(self.doc_font)
			self.color_square()
			self.comments_and_description(comments,description)
			self.actual_and_expected(actual,expected)
			err_fill = PatternFill(start_color='00FFFF00',
			                   end_color='0000FF00',
			                   fill_type='solid')
			self.results[self.stepcolumn+str(self.currentrow)]="step "+str(self.stepno)
			self.results[self.statuscolumn+str(self.currentrow)]="!ERROR!"
			self.results[self.statuscolumn+str(self.currentrow)].font=Font(name=self.document_font, bold=True, color="00FF0000")
			self.results[self.statuscolumn+str(self.currentrow)].fill=err_fill
			self.resultsbook.save(filename = self.bookname)
			self.stepno+=1
			self.currentrow+=1
		except Exception as e:
			err=e
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			#print(exc_type, fname, exc_tb.tb_lineno)
			##print("BHTest Failure: \nStep: "+str(self.stepno)+"\n"+description+"\n"+str(e)
			self.resultsbook.save(filename = self.bookname)
			self.results[self.stepcolumn+str(self.currentrow)]=" "+str(self.stepno)
			self.resultsbook.save(filename = self.bookname)

	def style_range(self, ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
	    top = Border(top=border.top)
	    left = Border(left=border.left)
	    right = Border(right=border.right)
	    bottom = Border(bottom=border.bottom)

	    first_cell = ws[cell_range.split(":")[0]]
	    if alignment:
	        ws.merge_cells(cell_range)
	        first_cell.alignment = alignment

	    rows = ws[cell_range]
	    if font:
	        first_cell.font = font

	    for cell in rows[0]:
	        cell.border = cell.border + top
	    for cell in rows[-1]:
	        cell.border = cell.border + bottom

	    for row in rows:
	        l = row[0]
	        r = row[-1]
	        l.border = l.border + left
	        r.border = r.border + right
	        if fill:
	            for c in row:
	                c.fill = fill

	def action_end(self):
		#passing=
		#failing=
		#time=
		#description = 
		self.color_square()
		thin = Side(border_style="thin", color="000000")
		#double = Side(border_style="double", color="ff0000")
		border = Border(top=thin, left=thin, right=thin)
		fill = PatternFill("solid", fgColor="FFFFFFFF")
		font = Font(size=12, b=False, color="000000", name=self.document_font)
		al = Alignment(horizontal="center", vertical="center", wrap_text=True)
		action_description=str(self.stepcolumn)+str((self.currentrow))

		self.style_range(self.results, "B"+str(self.currentrow)+':G'+str(self.currentrow), border, fill, font, al)
		self.results["B"+str(self.currentrow)]=self.action_description
		#Description is set

		self.currentrow+=1

		self.resultsbook.save(filename = self.bookname)


	def insert_action(self,title,description):
		if type(self.results['I1'].value)==None:
			self.results['I1']='1'
		try:
			h=self.results['I1'].value
			#print('try h')
			int(h)
			#print('try in h')
		except Exception as e:
			##print(type(e)
			#print(e.args

			self.results['I1']='1'
		self.action_description=description
		self.current_action=int(self.results['I1'].value)
		self.current_action_fail=0
		self.current_action_pass=0
		self.current_action+=1
		self.results['I1'].value=str(self.current_action)
		self.resultsbook.save(filename = self.bookname)
		self.action_name="Action "+str(self.current_action)+"\n"+title
		current_action=self.current_action
		title = "Action "+str(current_action)+": "+str(title)
		description = str(description)
		thin = Side(border_style="thin", color="000000")
		#double = Side(border_style="double", color="ff0000")
		border = Border(top=thin, left=thin, right=thin, bottom=thin)
		fill = PatternFill("solid", fgColor="DDDDDD")
		font = Font(name=self.document_font, size=18, b=True, color="000000")
		al = Alignment(horizontal="center", vertical="center", wrap_text=True)
		action_title=str(self.stepcolumn)+str(self.currentrow)

		self.color_square()
		self.style_range(self.results, "A"+str(self.currentrow)+':H'+str(self.currentrow), border, fill, font, al)
		self.results["A"+str(self.currentrow)]=title
		self.currentrow+=1

		linebreak=int(self.action_description_column_linebreak)
		description_length=len(description)
		self.results.row_dimensions[self.currentrow].height=int(15*description_length/linebreak+15)

		self.color_square()
		thin = Side(border_style="thin", color="000000")
		#double = Side(border_style="double", color="ff0000")
		border = Border(top=thin, left=thin, right=thin)
		fill = PatternFill("solid", fgColor="FFFFFFFF")
		font = Font(size=12, b=False, color="000000", name=self.document_font)
		al = Alignment(horizontal="center", vertical="center", wrap_text=True)
		action_description=str(self.stepcolumn)+str((self.currentrow))
		self.style_range(self.results, "B"+str(self.currentrow)+':G'+str(self.currentrow), border, fill, font, al)
		self.results["B"+str(self.currentrow)]=description
		description_length=len(str(description))

		#Description is set
		self.currentrow+=1
		#self.results[]=title
		#self.results[action_description]=description
		self.stepno=1
		self.resultsbook.save(filename = self.bookname)
		self.currentrow+=1
		currow=int(self.currentrow)
		self.print_titles(currow)
		self.resultsbook.save(filename = self.bookname)

class loadtest(nametest):
	def __init__(self, testname, title, description, path_to_results):
		self.path_to_results= path_to_results
		try:
			self.resultsbook=load_workbook(self.path_to_results)
		except:
			self.path_to_results=self.path_to_results.replace('\\','/')
			self.path_to_results=self.path_to_results.replace('//','/')
			self.path_to_results=self.path_to_results.replace('\\\\','/')
			folder_list=self.path_to_results.split('/')
			dir_string=''
			for directory in folder_list:
				#print(dir_string
				if len(dir_string)<1:
					dir_string=directory
				elif '.xlsx' in directory:
					wb=openpyxl.Workbook()
					wb.save(self.path_to_results)
				else:
					dir_string=dir_string+'/'+directory
					try:
						os.mkdir(dir_string)
					except:
						pass
		self.resultsbook=load_workbook(self.path_to_results)
		self.results = self.resultsbook.active
		self.tccolumn="C"
		self.startrow=0
		self.statuscolumn="G"
		self.arcolumn="F"
		self.ercolumn="E"
		self.cmntcolumn="H"
		self.desccolumn="D"
		self.stepcolumn="B"
		self.refcolumn="I"
		self.currentrow=self.results.max_row+2
		self.action_name=title
		self.bookname=path_to_results
		if type(self.results['I1'].value)!=str:
			self.results['I1']='0'
		self.current_action=int(self.results['I1'].value)
		if self.results['I6']!=str:
			self.results['I6']='0'
		self.total_action_pass=int(self.results['I6'].value)
		if self.results['I7']!=str:
			self.results['I7']='0'
		self.total_action_fail=int(self.results['I7'].value)
		#print(self.results['J6'].value
		if self.results['J6'].value is not None:
			self.total_pass=self.results['J6'].value
		else:
			self.total_pass=0
		if self.results['J7'].value is not None:
			self.total_fail=self.results['J7'].value
		else:
			self.total_fail=0
		print('results totalling pass')
		#print('check 1'
		self.stepno=1
		self.testname = testname
		self.passno=0
		self.bookname = path_to_results
		self.directory=path_to_results
		self.resultsbook.save(filename = self.bookname)
		print('results save')
		self.date=datetime.datetime.now().strftime("%m_%d")		
		#print('check 2'
		print('strftime')
		self.action_description_column_linebreak=131
		print('linebreak=131')

		self.square_fill = PatternFill(start_color='AAAAAAAA',
		                   end_color='AAAAAAAA',
		                   fill_type='solid')
		print('squarefill')
		self.bd = Side(style='thin', color="000000")
		self.document_font="Calisto MT"
		print('fontset')
		self.doc_font=Font(name = self.document_font, size =12)
		self.start_font=Font(name = self.document_font, bold=True, size = 18)
		print('start fill')
		self.start_fill = PatternFill(start_color='99999999', end_color='99999999',fill_type='solid')
		print('patternfill')
		print('title: '+title)
		print('description'+description)
		self.insert_action(str(title),str(description))
		print('insert action')
		#self.print_titles(self.startrow)
		#print('check 3'
		print ('function end')
class headertest(nametest):
	def __init__(self, testname, title, description, path_to_results, headers_dict):
		self.path_to_results= path_to_results
		self.headers_dict=headers_dict
		try:
			self.resultsbook=load_workbook(self.path_to_results)
		except:
			self.path_to_results=self.path_to_results.replace('\\','/')
			self.path_to_results=self.path_to_results.replace('//','/')
			self.path_to_results=self.path_to_results.replace('\\\\','/')
			folder_list=self.path_to_results.split('/')
			dir_string=''
			for directory in folder_list:
				#print(dir_string
				if len(dir_string)<1:
					dir_string=directory
				elif '.xlsx' in directory:
					wb=openpyxl.Workbook()
					wb.save(self.path_to_results)
				else:
					dir_string=dir_string+'/'+directory
					try:
						os.mkdir(dir_string)
					except:
						pass
		self.resultsbook=load_workbook(self.path_to_results)
		self.results = self.resultsbook.get_active_sheet()
		self.startrow=0

		self.tccolumn="C"
		self.statuscolumn="G"
		self.arcolumn="F"
		self.ercolumn="E"
		self.cmntcolumn="H"
		self.desccolumn="D"
		self.stepcolumn="B"
		self.refcolumn="I"



		self.currentrow=self.results.max_row+2
		self.action_name=title
		self.bookname=path_to_results
###

		if self.results['I1']!=str:
			self.results['I1']='0'
		self.current_action=int(self.results['I1'].value)

		if self.results['I6']!=str:
			self.results['I6']='0'
		self.total_action_pass=int(self.results['I6'].value)

		if self.results['I7']!=str:
			self.results['I7']='0'
		self.total_action_fail=int(self.results['I7'].value)

		#print(self.results['J6'].value
		if self.results['J6'].value is not None:
			self.total_pass=self.results['J6'].value
		else:
			self.total_pass=0
		if self.results['J7'].value is not None:
			self.total_fail=self.results['J7'].value
		else:
			self.total_fail=0
###

		#print('check 1'
		self.stepno=1
		self.testname = testname
		self.passno=0
		self.bookname = path_to_results
		self.directory=path_to_results
		self.resultsbook.save(filename = self.bookname)

		self.date=datetime.datetime.now().strftime("%m_%d")		
		#print('check 2'
		self.action_description_column_linebreak=131

		self.square_fill = PatternFill(start_color='AAAAAAAA',
		                   end_color='AAAAAAAA',
		                   fill_type='solid')
		self.bd = Side(style='thin', color="000000")
		self.document_font="Calisto MT"
		self.doc_font=Font(name = self.document_font, size =12)
		self.start_font=Font(name = self.document_font, bold=True, size = 18)
		self.start_fill = PatternFill(start_color='99999999', end_color='99999999',fill_type='solid')
#		self.insert_action(title,description)
		#print('check 3'
	def print_titles(self,row):
		#print(self.currentrow
		if self.currentrow>3:
			pass
		else:

			#print(type(self.results)
			#print(self.results['I6']

			self.currentrow+=1
			self.color_square()
			print('start titles1.1')

			# stepcolumn=self.results[self.stepcolumn+str(row)]
			# tccolumn=self.results[self.tccolumn+str(row)]
			# statuscolumn=self.results[self.statuscolumn+str(row)]
			# arcolumn=self.results[self.arcolumn+str(row)]
			# ercolumn=self.results[self.ercolumn+str(row)]
			# cmntcolumn=self.results[self.cmntcolumn+str(row)]
			# desccolumn=self.results[self.desccolumn+str(row)]



			# self.results[self.stepcolumn+str(row)]="Step"
			# self.results[self.tccolumn+str(row)]="Testcase"
			# self.results[self.statuscolumn+str(row)]="Status"
			# self.results[self.arcolumn+str(row)]="Actual Result"
			# self.results[self.ercolumn+str(row)]="Expected Result"
			# self.results[self.cmntcolumn+str(row)]="Comments        "
			# self.results[self.desccolumn+str(row)]="Description     "

			for key in self.headers_dict.keys():
				column_letter=self.headers_dict[key]
				self.results[str(column_letter)+str(row)]=str(key)
			print('start titles1.2')

			cell_list=[]
			for key in self.headers_dict.keys():
				column_letter=self.headers_dict[key]
				key_column=self.results[str(column_letter)+str(row)]
				#print('key: '+str(key)
				#print(key_column
				cell_list.append(key_column)
			print('start titles1.3')

			for cell in cell_list:
				cell='sploop'


			print('start titles1.4')

			# cell_list.append(stepcolumn)
			# #print(stepcolumn
			# cell_list.append(tccolumn)
			# cell_list.append(statuscolumn)
			# cell_list.append(arcolumn)
			# cell_list.append(ercolumn)
			# cell_list.append(cmntcolumn)
			# cell_list.append(desccolumn)
			for cell in cell_list:
				cell.font=self.start_font
				cell.fill=self.start_fill
				self.results.column_dimensions[cell.column].width=len(str(cell.value))*2.5
				cell.border = Border(left = self.bd, right = self.bd, bottom = self.bd, top = self.bd)
			print('start titles1.5')

	def step_fail(self,result_dict):
		try:
			self.format_row(self.doc_font)
			self.color_square()
#			self.comments_and_description(comments,description)
#			self.actual_and_expected(actual,expected)
			fail_fill = PatternFill(start_color='00FF0000',
			                   end_color='00FF0000',
			                   fill_type='solid')
			h_dict=self.headers_dict
			for item in h_dict:
				cell_col=str(h_dict[item])
				cell_row=str(self.currentrow)
				cell=self.results[cell_col+cell_row]
				if item in result_dict:
					cell.value=result_dict[item]
				else:
					if item=='Step':
						cell.value=str(self.stepno)
					if item=='Status':
						cell.value='Fail'
						cell.fill=fail_fill
					pass

			# self.results[self.stepcolumn+str(self.currentrow)]="step "+str(self.stepno)
			# self.results[self.tccolumn+str(self.currentrow)]=self.action_name

			# self.results[self.statuscolumn+str(self.currentrow)]="Fail"
			# self.results[self.statuscolumn+str(self.currentrow)].fill=fail_fill	

			self.resultsbook.save(filename = self.bookname)
			self.stepno+=1
			self.currentrow+=1
			self.total_action_fail+=1
			self.total_fail+=1
			self.results['I7']=self.total_action_fail
			self.results['J7']=self.total_fail
#			self.current_action_fail+=1
		except Exception as e:
			self.step_error(description, expected, actual,e)	

	def step_pass(self, result_dict):
		try:
			self.format_row(self.doc_font)
			self.color_square()
#			self.comments_and_description(comments,description)
#			self.actual_and_expected(actual,expected)

			pass_fill = PatternFill(start_color='5500FF00',
			                   end_color='5500FF00',
			                   fill_type='solid')
			h_dict=self.headers_dict
			for item in h_dict:
				cell_col=str(h_dict[item])
				cell_row=str(self.currentrow)
				cell=self.results[cell_col+cell_row]
				# #print(cell
				# #print(type(cell)
				# #print(dir(cell)
				# #print(cell.value
				# cell.value="bloop"
				# #print('post bloop:'
				# #print(cell.value
				if item in result_dict:
					cell.value=result_dict[item]
				else:
					if item=='Step':
						cell.value=str(self.stepno)
					if item=='Status':
						cell.value='Pass'
						cell.fill=pass_fill
					pass

			self.resultsbook.save(filename = self.bookname)
			self.stepno+=1
			self.currentrow+=1
			self.total_action_pass+=1
			self.total_pass+=1
			self.results['I6']=self.total_action_pass
			self.results['J6']=self.total_pass
#			self.current_action_pass+=1
		except Exception as e:
			print(e)
#			self.step_error(description, expected, actual,e)	
class Databook:
#1/31/2017 - Changed the variable name "self.ws" to self.curent_sheet
	def __init__(self,filelocation):
		self.vertical=False
		self.filelocation=filelocation
		self.workbook=load_workbook(filelocation)
		self.current_sheet = self.workbook.active
		self.columns=self.current_sheet.max_column
		self.rows=self.current_sheet.max_row

	def currentsheet(self):
		item = str(self.current_sheet)
		item=item.strip('<Worksheet "')
		item=item.strip('">')
		return item

	def load_sheet(self,sheet_name):
		self.current_sheet=self.workbook.get_sheet_by_name(sheet_name)
		if self.current_sheet['A1'].value=='vertical':
			self.vertical=True
		else:
			self.vertical=False
		self.columns=self.current_sheet.max_column
		self.rows=self.current_sheet.max_row
		
	def fetch_data(self,row_num):
		if self.vertical==True:
			dataList=[]
			#rows
			data_rows=self.current_sheet.rows
			for cellObj in data_rows:
				appendit=cellObj[row_num].value
				dataList.append(appendit)
			return dataList
		else:
			dataList=[]
			columns=self.current_sheet.columns
			for cellObj in columns:
				appendit=cellObj[row_num].value
				dataList.append(appendit)
			return dataList
	def get_titles(self):
		if self.vertical==True:
			return self.fetch_data(0)
		return self.fetch_data(1)	
	def fetch_dictionary(self):
		print(self.current_sheet)
		if self.vertical==True:
			dictlist=[]
			#max_row=self.current_sheet.max_row
			max_column=self.current_sheet.max_column
			symbols=['-','\\','|','/']
			mod=1
			for y in range (2,max_column):
				percentage=str(y*100/max_column)
				symbols=['-','\\','|','/']
				mod=y%4
				sys.stdout.write('\r'+symbols[mod]+' %'+percentage+' complete\trow num:'+str(y))
				
				datadic={}
				titles=self.get_titles()
				data=self.fetch_data(y)			
				for x in range (0,self.rows):
					datadic.update({titles[x]:str(data[x])})
				dictlist.append(datadic)
			sys.stdout.write('\r'+symbols[mod]+' %100 complete')
			sys.stdout.flush()
			print('self.current_sheet')
			return dictlist
		else:
			dictlist=[]
			max_row=self.current_sheet.max_row
			symbols=['-','\\','|','/']
			mod=1
			for y in range (2,max_row):
				percentage=str(y*100/max_row)
				symbols=['-','\\','|','/']
				mod=y%4
				sys.stdout.write('\r'+symbols[mod]+' %'+percentage+' complete\trow num:'+str(y))
				
				datadic={}
				titles=self.get_titles()
				data=self.fetch_data(y)			
				for x in range (0,self.columns):
					datadic.update({titles[x]:str(data[x])})
				dictlist.append(datadic)
			sys.stdout.write('\r'+symbols[mod]+' %100 complete')
			sys.stdout.flush()
			return dictlist
	def get_actions(self):
		actionslist=self.workbook.get_sheet_names()
		return actionslist